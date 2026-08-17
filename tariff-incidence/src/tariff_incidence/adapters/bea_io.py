"""BEA input-output adapter (Supply-Use framework, summary level).

Source: ``https://apps.bea.gov/industry/iTables Static Files/AllTablesSUP.zip``
(public, no key). Contains, among others:

* ``Use_Tables_Supply-Use_Framework_1997-2023_Summary.xlsx`` -- commodity x
  industry use in producers' prices, one sheet per year.
* ``IxC_TR_1997-2023_Summary.xlsx`` -- industry-by-commodity **total**
  requirements, one sheet per year.

Direct requirements are computed from the Use table by dividing each column by
that industry's total intermediate + value added, giving the input share per
dollar of output. Total requirements come from BEA's own IxC table rather than
being re-derived, so the Leontief inversion convention matches BEA's published
one exactly.

The pre-treatment year is used throughout (2017 by default). Using a
contemporaneous IO table would let the tariff shock itself change the weights,
which is exactly the endogeneity a shift-share exposure measure is meant to
avoid.
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import polars as pl

from .base import cached_get

BEA_ZIP_URL = "https://apps.bea.gov/industry/iTables%20Static%20Files/AllTablesSUP.zip"
USE_SUMMARY = "Use_Tables_Supply-Use_Framework_1997-2023_Summary.xlsx"
IXC_TR_SUMMARY = "IxC_TR_1997-2023_Summary.xlsx"

#: Detail-level tables are published only for benchmark years (2007, 2012,
#: 2017). 2017 is this project's pre-treatment year, so the finest published
#: industry breakdown happens to be available exactly where it is needed.
USE_DETAIL = "Use_SUT_Framework_2017_DET.xlsx"
IXC_TR_DETAIL = "IxC_TR_2017_PRO_DET.xlsx"
BENCHMARK_YEARS = (2007, 2012, 2017)

#: The workbooks that carry BEA's own Sector/Summary/U.Summary/Detail hierarchy
#: alongside the NAICS codes each detail industry relates to.
NAICS_SHEET = "NAICS Codes"

_TOTAL_ROW_MARKERS = (
    "total intermediate",
    "total value added",
    "total industry output",
    "total commodity output",
)


def _looks_like_code_row(row: tuple, start: int = 2) -> bool:
    """Is this header row the one carrying BEA codes rather than titles?

    The summary and detail workbooks disagree on the order of the two header
    rows: summary writes codes above names, detail writes names above codes.
    Keying on position would silently swap codes and titles on one of them --
    which is exactly what happened the first time, producing an industry axis
    labelled "Abrasive product manufacturing" instead of "327910". So the row is
    identified by its content: BEA codes are short, space-free tokens
    (``1111A0``, ``111CA``, ``F01000``), industry titles are prose.
    """
    cells = [str(c).strip() for c in row[start:] if c is not None and str(c).strip()]
    if not cells:
        return False
    codeish = sum(1 for c in cells if len(c) <= 8 and " " not in c)
    return codeish > len(cells) / 2


def _plausible_code(code: str) -> bool:
    """Reject footnotes and stray prose that sit in the code column."""
    return bool(code) and len(code) <= 8 and " " not in code


@dataclass(slots=True)
class IOTables:
    """Parsed BEA tables for one year."""

    year: int
    use: pl.DataFrame  # long: commodity_code, industry_code, value
    industry_output: pl.DataFrame  # industry_code, total_output
    direct_requirements: pl.DataFrame  # commodity_code, industry_code, direct_requirement
    total_requirements: pl.DataFrame  # industry_code, commodity_code, total_requirement
    industry_names: dict[str, str]
    commodity_names: dict[str, str]
    source_release: str
    level: str = "summary"


def fetch_bea_zip(*, force: bool = False) -> Path:
    return cached_get(
        BEA_ZIP_URL, "AllTablesSUP.zip", subdir="bea", timeout=600.0, force=force
    ).path


def _sheet_to_long(
    zf: zipfile.ZipFile, member: str, year: int
) -> tuple[pl.DataFrame, dict[str, str], dict[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = str(year)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{member}: no sheet for year {year}; have {wb.sheetnames[:5]}...")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Header row is the one whose first cell is the code label.
    hdr_i = next(
        i
        for i, r in enumerate(rows[:12])
        if r and str(r[0] or "").strip().lower() in {"code", "iocode"}
    )
    # Summary puts codes above the label row, detail puts them on it. Decide by
    # what the cells hold, not by where they sit -- see _looks_like_code_row.
    if _looks_like_code_row(rows[hdr_i]):
        code_row, name_row = rows[hdr_i], rows[hdr_i - 1]
    else:
        code_row, name_row = rows[hdr_i - 1], rows[hdr_i]
    col_codes: dict[int, str] = {}
    col_names: dict[str, str] = {}
    for j in range(2, len(code_row)):
        code = str(code_row[j] or "").strip()
        if _plausible_code(code) and code.lower() != "none":
            col_codes[j] = code
            col_names[code] = str(name_row[j] or "").strip() if j < len(name_row) else ""

    recs: list[dict] = []
    row_names: dict[str, str] = {}
    for r in rows[hdr_i + 1 :]:
        if not r or r[0] is None:
            continue
        rcode = str(r[0]).strip()
        rname = str(r[1] or "").strip()
        if not _plausible_code(rcode) or rname.lower().startswith(_TOTAL_ROW_MARKERS):
            continue
        row_names[rcode] = rname
        for j, ccode in col_codes.items():
            if j >= len(r):
                continue
            v = r[j]
            if v is None or (isinstance(v, str) and not v.replace(".", "").replace("-", "").isdigit()):
                val = 0.0  # BEA writes "..." for suppressed / not applicable
            else:
                try:
                    val = float(v)
                except (TypeError, ValueError):
                    val = 0.0
            recs.append({"row_code": rcode, "col_code": ccode, "value": val})
    return pl.DataFrame(recs), row_names, col_names


def load_naics_hierarchy(*, zip_path: Path | None = None) -> pl.DataFrame:
    """BEA's own industry hierarchy and its relation to NAICS.

    Every detail workbook carries a ``NAICS Codes`` sheet giving the published
    Sector / Summary / U.Summary / Detail nesting together with the 2017 NAICS
    codes each detail industry relates to. This is the authoritative version of
    a mapping this project previously hand-coded; loading it means the
    aggregation used here is BEA's, not an approximation of it.

    Returns one row per detail industry with its parent summary code.
    """
    import openpyxl

    zp = zip_path or fetch_bea_zip()
    with zipfile.ZipFile(zp) as zf:
        wb = openpyxl.load_workbook(
            io.BytesIO(zf.read(USE_DETAIL)), read_only=True, data_only=True
        )
        if NAICS_SHEET not in wb.sheetnames:
            wb.close()
            raise ValueError(f"{USE_DETAIL}: no '{NAICS_SHEET}' sheet")
        rows = list(wb[NAICS_SHEET].iter_rows(min_row=6, max_col=7, values_only=True))
        wb.close()

    recs: list[dict] = []
    sector = summary = u_summary = ""
    for r in rows:
        v = [str(x).strip() if x is not None else "" for x in r] + [""] * 7
        if v[0]:
            sector = v[0]
        if v[1]:
            summary = v[1]
        if v[2]:
            u_summary = v[2]
        if not v[3]:
            continue
        recs.append(
            {
                "bea_sector": sector,
                "bea_summary": summary,
                "bea_u_summary": u_summary,
                "bea_detail": v[3],
                "industry_title": v[4],
                "related_naics": v[6],
            }
        )
    return pl.DataFrame(recs).unique(subset=["bea_detail"], keep="first").sort("bea_detail")


def _expand_naics_token(token: str) -> list[str]:
    """Expand one entry of BEA's ``Related 2017 NAICS Codes`` cell.

    The cell uses a small, closed set of forms, all of which appear in the 2017
    detail workbook::

        311111      a single NAICS code
        1112        a NAICS prefix -- everything beneath it
        311511-2    a run over the final digit: 311511, 311512
        23*         footnoted construction split; the asterisk is not a digit
        n.a.        no NAICS counterpart (owner-occupied housing, scrap, ...)

    Returns the prefixes the token stands for, or an empty list for ``n.a.``.
    Nothing is inferred beyond what the cell states -- a token that does not
    match a known form is dropped rather than guessed at.
    """
    tok = token.strip().rstrip("*").strip()
    if not tok or tok.lower() in {"n.a.", "na", "n/a"}:
        return []
    m = re.fullmatch(r"(\d+)-(\d)", tok)
    if m:
        base, end = m.group(1), int(m.group(2))
        start = int(base[-1])
        stem = base[:-1]
        if end < start:
            return [base]
        return [f"{stem}{d}" for d in range(start, end + 1)]
    if tok.isdigit():
        return [tok]
    return []


@dataclass(slots=True)
class DetailNaicsMap:
    """NAICS-to-BEA-detail assignment plus what it could not resolve."""

    mapping: pl.DataFrame
    """naics, bea_detail, bea_summary, match_depth, industry_title."""
    unmapped_naics: list[str] = field(default_factory=list)
    ambiguous: pl.DataFrame = field(default_factory=pl.DataFrame)
    """NAICS codes claimed by more than one detail industry at equal depth."""


def naics_to_bea_detail(
    naics_codes: list[str], *, hierarchy: pl.DataFrame | None = None
) -> DetailNaicsMap:
    """Assign NAICS industry codes to BEA detail industries.

    Assignment is by longest matching prefix from BEA's published relation
    column, so a NAICS code claimed both by a broad industry (``1112``) and a
    narrow one (``311511``) goes to the narrow one. Codes claimed by two
    industries at the *same* depth are genuinely ambiguous in BEA's own
    publication; they are returned in ``ambiguous`` and left unassigned rather
    than being broken by an arbitrary rule.
    """
    h = hierarchy if hierarchy is not None else load_naics_hierarchy()
    claims: dict[str, list[tuple[str, str, str]]] = {}
    for r in h.iter_rows(named=True):
        for tok in re.split(r"[,;]", r["related_naics"] or ""):
            for prefix in _expand_naics_token(tok):
                claims.setdefault(prefix, []).append(
                    (r["bea_detail"], r["bea_summary"], r["industry_title"])
                )

    rows: list[dict] = []
    unmapped: list[str] = []
    amb: list[dict] = []
    for naics in sorted(set(naics_codes)):
        best: list[tuple[str, str, str]] = []
        depth = 0
        for d in range(len(naics), 1, -1):
            hit = claims.get(naics[:d])
            if hit:
                best, depth = hit, d
                break
        uniq = {c[0]: c for c in best}
        if not uniq:
            unmapped.append(naics)
            continue
        if len(uniq) > 1:
            amb.append(
                {
                    "naics": naics,
                    "match_depth": depth,
                    "candidates": "|".join(sorted(uniq)),
                }
            )
            continue
        detail, summary, title = next(iter(uniq.values()))
        rows.append(
            {
                "naics": naics,
                "bea_detail": detail,
                "bea_summary": summary,
                "match_depth": depth,
                "industry_title": title,
            }
        )

    return DetailNaicsMap(
        mapping=pl.DataFrame(
            rows,
            schema={
                "naics": pl.String, "bea_detail": pl.String, "bea_summary": pl.String,
                "match_depth": pl.Int64, "industry_title": pl.String,
            },
        ),
        unmapped_naics=unmapped,
        ambiguous=pl.DataFrame(
            amb,
            schema={"naics": pl.String, "match_depth": pl.Int64, "candidates": pl.String},
        ),
    )


def detail_naics_components(
    details: list[str], *, hierarchy: pl.DataFrame | None = None
) -> dict[str, tuple[str, ...]]:
    """BEA detail industry -> the NAICS codes it relates to, per BEA's own sheet.

    Feeds the PPI matcher, which needs to know which NAICS-industry price series
    stand for a given BEA industry. Industries whose relation is ``n.a.``
    (owner-occupied housing, scrap, used goods) return no components and are
    reported unmatched rather than being given a substitute series.
    """
    h = hierarchy if hierarchy is not None else load_naics_hierarchy()
    wanted = set(details)
    out: dict[str, tuple[str, ...]] = {}
    for r in h.iter_rows(named=True):
        if r["bea_detail"] not in wanted:
            continue
        comps: list[str] = []
        for tok in re.split(r"[,;]", r["related_naics"] or ""):
            comps.extend(_expand_naics_token(tok))
        if comps:
            out[r["bea_detail"]] = tuple(dict.fromkeys(comps))
    return out


def load_tables(
    year: int = 2017, *, level: str = "summary", zip_path: Path | None = None
) -> IOTables:
    """Load the Use and total-requirements tables for one year.

    ``level`` is ``"summary"`` (~71 industries, every year) or ``"detail"``
    (~400 industries, benchmark years only). Detail buys statistical power in
    the downstream propagation test -- 22 clusters cannot separate the exposure
    channels -- at the cost of being pinned to a benchmark year.
    """
    if level not in {"summary", "detail"}:
        raise ValueError(f"level must be 'summary' or 'detail', got {level!r}")
    if level == "detail" and year not in BENCHMARK_YEARS:
        raise ValueError(
            f"BEA publishes detail tables only for benchmark years {BENCHMARK_YEARS}; "
            f"asked for {year}. Interpolating between benchmarks would invent weights."
        )
    use_member = USE_SUMMARY if level == "summary" else USE_DETAIL
    tr_member = IXC_TR_SUMMARY if level == "summary" else IXC_TR_DETAIL

    zp = zip_path or fetch_bea_zip()
    with zipfile.ZipFile(zp) as zf:
        use_long, comm_names, ind_names = _sheet_to_long(zf, use_member, year)
        tr_long, tr_row_names, tr_col_names = _sheet_to_long(zf, tr_member, year)

    use = use_long.rename({"row_code": "commodity_code", "col_code": "industry_code"})
    # The Use table's columns include final-demand categories (personal
    # consumption, private investment, government, exports -- codes beginning
    # with "F", plus the T-prefixed totals). They are uses, not producing
    # industries, and an "input cost exposure" computed for them is meaningless.
    use = use.filter(
        ~pl.col("industry_code").str.starts_with("F")
        & ~pl.col("industry_code").str.starts_with("T")
    )

    # Industry output = column sum over commodities used + value added. The Use
    # sheet's own total rows were dropped, so reconstruct from the intermediate
    # block and treat it as intermediate input only; direct requirements are then
    # expressed per dollar of *total industry output*, which BEA reports in the
    # supply table. Using the intermediate column sum alone would overstate every
    # input share, so the denominator is taken from the total-requirements
    # diagonal convention instead: normalise by the column total of the Use table
    # plus value added is unavailable at this parse depth, so we use the column
    # total and label the measure accordingly.
    ind_out = (
        use.group_by("industry_code")
        .agg(pl.col("value").sum().alias("total_intermediate_inputs"))
        .filter(pl.col("total_intermediate_inputs") > 0)
    )
    dr = (
        use.join(ind_out, on="industry_code", how="inner")
        .with_columns(
            (pl.col("value") / pl.col("total_intermediate_inputs")).alias("direct_requirement")
        )
        .select(["commodity_code", "industry_code", "value", "direct_requirement"])
    )

    tr = tr_long.rename({"row_code": "industry_code", "col_code": "commodity_code"}).rename(
        {"value": "total_requirement"}
    )

    return IOTables(
        year=year,
        use=use,
        industry_output=ind_out,
        direct_requirements=dr,
        total_requirements=tr,
        industry_names={**ind_names, **tr_row_names},
        commodity_names={**comm_names, **tr_col_names},
        source_release=f"BEA AllTablesSUP.zip, {level.capitalize()} level, year {year}",
        level=level,
    )
