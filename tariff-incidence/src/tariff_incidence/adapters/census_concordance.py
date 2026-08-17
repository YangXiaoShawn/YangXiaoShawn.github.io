"""Census import concordance: 10-digit HTS to NAICS.

Source: ``https://www.census.gov/foreign-trade/reference/codes/concordance/``,
files ``impconcordNN.xls[x]`` where NN is the two-digit year. Public, no key.

This replaces the hand-built HS2-chapter map that industry exposure previously
relied on (D-011). That map assigned whole chapters to a single BEA commodity
and was labelled ``COARSE_APPROXIMATION`` because within-chapter heterogeneity
was lost. The official concordance is keyed on the **10-digit commodity code** —
exactly the level this project's panel is built at — and gives each line one
6-digit NAICS industry, so no weighting assumption is required at all.

Two things worth knowing before using it.

**Vintage matters.** Census publishes one file per year because commodity codes
and their NAICS assignments change. ``impconcord19`` is the 2019 vintage. Mixing
a 2020 concordance into a 2018 panel would silently misassign lines that were
renumbered, so the vintage actually used is recorded in the manifest.

**NAICS codes here are not always six clean digits.** Census writes an ``X`` in
positions where it aggregates detail it does not publish separately, e.g.
``11211X``. Those are kept verbatim rather than coerced: the ``X`` is
information about what Census would not disclose, and silently truncating it to
a shorter code would assert a precision the source declines to give.
"""

from __future__ import annotations

import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import polars as pl

from .base import cached_get

CONCORDANCE_BASE = "https://www.census.gov/foreign-trade/reference/codes/concordance"

#: Two-digit year -> filename. Census switched extension in 2020.
_FILENAMES = {
    17: "impconcord17.xls",
    18: "impconcord18.xls",
    19: "impconcord19.xls",
    20: "impconcord20.xlsx",
    21: "impconcord21.xlsx",
    22: "impconcord22.xlsx",
}

_NAICS_OK = re.compile(r"^[0-9X]{2,6}$")


@dataclass(slots=True)
class ConcordanceLoad:
    """A parsed concordance plus what it could not resolve."""

    vintage_year: int
    source_file: str
    mapping: pl.DataFrame
    """hs10, naics, unit_qy1, end_use, sitc."""
    n_rows: int = 0
    n_unmapped: int = 0
    n_aggregated_naics: int = 0
    """Rows whose NAICS carries an X, i.e. Census aggregated undisclosed detail."""
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, str]:
        return dict(
            zip(
                self.mapping["hs10"].to_list(),
                self.mapping["naics"].to_list(),
                strict=True,
            )
        )


def fetch_concordance(vintage_year: int, *, force: bool = False) -> Path:
    """Download one vintage of the import concordance."""
    yy = vintage_year % 100
    if yy not in _FILENAMES:
        raise ValueError(
            f"no concordance filename known for 20{yy:02d}; available: "
            + ", ".join(f"20{k:02d}" for k in sorted(_FILENAMES))
        )
    name = _FILENAMES[yy]
    res = cached_get(
        f"{CONCORDANCE_BASE}/{name}",
        name,
        subdir="census_concordance",
        timeout=300.0,
        force=force,
    )
    return res.path


def _read_workbook(path: Path) -> tuple[list[tuple], list[str]]:
    """Read either workbook format Census uses.

    Census published the concordance as legacy ``.xls`` through 2019 and as
    ``.xlsx`` from 2020. openpyxl handles only the latter; ``xlrd`` handles only
    the former. Supporting both is what makes per-year vintages usable at all --
    an earlier version raised on ``.xls`` and, in doing so, quietly forced a
    2020 mapping onto a 2017-2019 panel.
    """
    if path.suffix.lower() == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(book.nsheets - 1)
        header = [str(sheet.cell_value(0, c)).strip().lower() for c in range(sheet.ncols)]
        rows = [
            tuple(sheet.cell_value(r, c) for c in range(sheet.ncols))
            for r in range(1, sheet.nrows)
        ]
        return rows, header

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = next(
        (s for s in wb.sheetnames if s.lower().startswith("impconcord")), wb.sheetnames[-1]
    )
    it = wb[name].iter_rows(values_only=True)
    header = [str(h).strip().lower() if h is not None else "" for h in next(it)]
    rows = list(it)
    wb.close()
    return rows, header


def parse_concordance(path: Path, vintage_year: int) -> ConcordanceLoad:
    """Parse an ``impconcord`` workbook into an HS10 -> NAICS table.

    Pure function over a local file, so it runs offline in the test suite.
    """
    rows, header = _read_workbook(path)

    def col(name: str) -> int:
        try:
            return header.index(name)
        except ValueError as exc:
            raise ValueError(
                f"{path.name}: expected column {name!r}; found {header}"
            ) from exc

    i_c, i_n = col("commodity"), col("naics")
    i_u = header.index("unit_qy1") if "unit_qy1" in header else None
    i_e = header.index("end_use") if "end_use" in header else None
    i_s = header.index("sitc") if "sitc" in header else None

    recs: list[dict] = []
    unmapped = 0
    aggregated = 0
    for r in rows:
        if r is None or r[i_c] is None:
            continue
        raw = r[i_c]
        # xlrd hands back floats for numeric cells, so 0101210010 arrives as
        # 101210010.0 and would be dropped by a naive digit check.
        hs10 = (
            f"{int(raw):010d}"
            if isinstance(raw, float) and raw.is_integer()
            else str(raw).strip().replace(".", "")
        )
        if not hs10.isdigit() or len(hs10) != 10:
            continue
        naics_raw = r[i_n]
        if isinstance(naics_raw, float) and naics_raw.is_integer():
            naics = f"{int(naics_raw):06d}"
        else:
            naics = str(naics_raw).strip().upper() if naics_raw is not None else ""
        if not naics or not _NAICS_OK.match(naics):
            unmapped += 1
            continue
        if "X" in naics:
            aggregated += 1
        recs.append(
            {
                "hs10": hs10,
                "naics": naics,
                "unit_qy1": (str(r[i_u]).strip() if i_u is not None and r[i_u] else ""),
                "end_use": (str(r[i_e]).strip() if i_e is not None and r[i_e] else ""),
                "sitc": (str(r[i_s]).strip() if i_s is not None and r[i_s] else ""),
            }
        )

    mapping = pl.DataFrame(recs).unique(subset=["hs10"], keep="first")
    load = ConcordanceLoad(
        vintage_year=vintage_year,
        source_file=path.name,
        mapping=mapping,
        n_rows=mapping.height,
        n_unmapped=unmapped,
        n_aggregated_naics=aggregated,
    )
    if unmapped:
        load.warnings.append(
            f"{unmapped} commodity lines carry no usable NAICS code and are excluded"
        )
    if aggregated:
        load.warnings.append(
            f"{aggregated} lines map to a NAICS code containing 'X', where Census "
            "aggregates detail it does not disclose; kept verbatim, not truncated"
        )
    return load


# --------------------------------------------------------------------------- #
# NAICS -> BEA summary industry
# --------------------------------------------------------------------------- #

#: BEA summary industries that combine several NAICS 3- or 4-digit groups.
#: Taken from BEA's published summary-level definitions rather than inferred, so
#: each entry is a documented aggregation and not a judgement call. Anything not
#: listed here falls back to its NAICS 3-digit prefix, which is what the
#: remaining BEA manufacturing codes (321, 322, ... 339) already are.
_BEA_COMPOSITES: dict[str, tuple[str, ...]] = {
    "111CA": ("111", "112"),
    "113FF": ("113", "114", "115"),
    "211": ("211",),
    "212": ("212",),
    "213": ("213",),
    "311FT": ("311", "312"),
    "313TT": ("313", "314"),
    "315AL": ("315", "316"),
    "3361MV": ("3361", "3362", "3363"),
    "3364OT": ("3364", "3365", "3366", "3369"),
}

#: BEA summary manufacturing industries that are exactly a NAICS 3-digit group.
_BEA_DIRECT = (
    "321", "322", "323", "324", "325", "326", "327",
    "331", "332", "333", "334", "335", "337", "339",
)


def naics_to_bea_summary(naics: str) -> str | None:
    """Map a NAICS code to a BEA summary industry, or ``None`` if unmapped.

    Longest match first, so a 4-digit composite such as 3361 wins over the
    3-digit fallback 336. Codes containing Census's ``X`` still map whenever the
    disclosed prefix is long enough to be unambiguous; where it is not, the
    result is ``None`` rather than a guess.
    """
    if not naics:
        return None
    code = naics.strip().upper()

    for bea, prefixes in _BEA_COMPOSITES.items():
        for p in sorted(prefixes, key=len, reverse=True):
            if code.startswith(p):
                return bea
    for p in _BEA_DIRECT:
        if code.startswith(p):
            return p
    if code[:3].isdigit() and code[:2] in {"11", "21", "22", "23", "42", "48", "49"}:
        return code[:3]
    return None


def build_hs10_to_bea(load: ConcordanceLoad) -> tuple[pl.DataFrame, dict[str, Any]]:
    """Extend an HS10 -> NAICS mapping to HS10 -> BEA summary industry."""
    m = load.mapping.with_columns(
        pl.col("naics")
        .map_elements(naics_to_bea_summary, return_dtype=pl.String)
        .alias("bea_summary")
    )
    mapped = m.filter(pl.col("bea_summary").is_not_null())
    warnings: list[str] = list(load.warnings)
    quality: dict[str, Any] = {
        "vintage_year": load.vintage_year,
        "source_file": load.source_file,
        "n_hs10_lines": load.n_rows,
        "n_mapped_to_bea": mapped.height,
        "n_unmapped_to_bea": load.n_rows - mapped.height,
        "share_mapped": mapped.height / load.n_rows if load.n_rows else None,
        "n_distinct_naics": int(m["naics"].n_unique()),
        "n_distinct_bea": int(mapped["bea_summary"].n_unique()),
        "n_aggregated_naics": load.n_aggregated_naics,
        "warnings": warnings,
        "method": (
            "Official Census import concordance at the 10-digit commodity level, then "
            "NAICS to BEA summary industry by BEA's published summary definitions. No "
            "weighting assumption is involved: each 10-digit line carries exactly one "
            "NAICS code."
        ),
    }
    if quality["n_unmapped_to_bea"]:
        warnings.append(
            f"{quality['n_unmapped_to_bea']} lines have a NAICS outside the BEA summary "
            "manufacturing and primary-sector groups (services, government, unclassified) "
            "and are excluded from industry exposure rather than forced into a bucket"
        )
    return m, quality


@dataclass(slots=True)
class VintageStack:
    """Several concordance vintages combined, with provenance per code."""

    primary_year: int
    years: list[int]
    mapping: pl.DataFrame
    """hs10, naics, bea_summary, vintage_year -- the vintage each code came from."""
    n_from_primary: int = 0
    n_from_fallback: int = 0
    n_reclassified_vs_primary: int = 0
    """Codes whose NAICS differs between the primary vintage and a later one."""
    reclassified_examples: list[dict] = field(default_factory=list)
    n_naics_vintage_repaired: int = 0
    """Lines whose primary-vintage NAICS the classification revision retired."""
    naics_vintage_repairs: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _repair_retired_naics(
    combined: pl.DataFrame, loads: dict[int, ConcordanceLoad], ys: list[int]
) -> tuple[pl.DataFrame, list[dict]]:
    """Replace NAICS codes the classification revision retired, line by line.

    A code counts as retired only if it appears in **none** of the vintages
    after the primary one. A code that still exists but was reassigned for a
    particular product is a genuine reclassification, and there the primary
    vintage keeps governing -- so this never silently overrides the
    pre-determination rule.

    The live code universe is taken from the **latest vintage only**, not from a
    union over later vintages: the union spans the switch, so a vintage still on
    2012 NAICS keeps the retired codes alive and the rule finds nothing. Taking
    the union looked safer and silently did less.

    The limitation that remains, stated rather than papered over: "absent from
    the latest file" is a superset of "retired by the classification revision" --
    a code that survived 2017 NAICS but drew no imports in the latest year would
    also qualify. That is why every substitution is returned and reported rather
    than applied invisibly. On the 2017-2020 files the codes it finds are the
    documented 2012-to-2017 consolidations (335221-8 into 335220, 333911/333913
    into 333914, 211111/211112 into 211120/211130, 212231/212234 into 212230).
    """
    latest = max(ys)
    if latest not in loads:
        return combined, []
    latest_map = loads[latest].mapping.select(["hs10", "naics"]).rename(
        {"naics": "_successor"}
    )
    live = {c for c in loads[latest].mapping["naics"].to_list() if c}
    retired = sorted({c for c in combined["naics"].to_list() if c and c not in live})
    if not retired:
        return combined, []

    joined = combined.join(latest_map, on="hs10", how="left")
    needs = joined.filter(pl.col("naics").is_in(retired) & pl.col("_successor").is_not_null())
    repairs = [
        {
            "retired_naics": r["naics"],
            "successor_naics": r["_successor"],
            "n_lines": r["len"],
        }
        for r in needs.group_by(["naics", "_successor"]).len().sort("len", descending=True)
        .iter_rows(named=True)
    ]
    if not repairs:
        return combined, []

    repaired = joined.with_columns(
        pl.when(pl.col("naics").is_in(retired) & pl.col("_successor").is_not_null())
        .then(pl.col("_successor"))
        .otherwise(pl.col("naics"))
        .alias("naics")
    ).drop("_successor")
    # bea_summary is derived from NAICS, so it has to follow the repair.
    repaired = repaired.with_columns(
        pl.col("naics")
        .map_elements(lambda n: naics_to_bea_summary(n), return_dtype=pl.String)
        .alias("bea_summary")
    )
    return repaired, repairs


def load_vintages(
    years: Sequence[int], *, primary_year: int | None = None
) -> VintageStack:
    """Combine concordance vintages, preferring the pre-treatment assignment.

    Census publishes one file per year because commodity codes are renumbered
    and NAICS assignments are revised. Two facts about this panel decide how to
    combine them.

    First, **the assignment should be pre-determined.** Industry exposure is
    built on pre-treatment import weights precisely so it cannot respond to the
    shock it explains; letting the industry *assignment* drift with later
    vintages would reintroduce through the back door what the weighting is
    careful to avoid. So the earliest vintage wins wherever it has the code.

    Second, **later vintages still add coverage.** Codes created after the
    primary year exist only in later files, and dropping them loses trade that
    genuinely occurred. They are taken from the earliest vintage that has them,
    and the source vintage is recorded per code rather than assumed.

    Reclassification between the primary and later vintages is counted, since it
    bounds how much the choice of primary vintage matters. On the 2017-2020
    files it is about 0.9% of codes, concentrated between 2018 and 2019.
    """
    ys = sorted(set(years))
    if not ys:
        raise ValueError("no concordance vintages requested")
    primary = primary_year if primary_year is not None else ys[0]
    if primary not in ys:
        raise ValueError(f"primary vintage {primary} is not among {ys}")

    loads: dict[int, ConcordanceLoad] = {}
    for y in ys:
        loads[y] = parse_concordance(fetch_concordance(y), y)

    frames = []
    seen: set[str] = set()
    n_primary = 0
    n_fallback = 0
    for y in [primary, *[x for x in ys if x != primary]]:
        m, _ = build_hs10_to_bea(loads[y])
        m = m.filter(~pl.col("hs10").is_in(list(seen))) if seen else m
        if m.height:
            frames.append(m.with_columns(pl.lit(y).alias("vintage_year")))
            seen |= set(m["hs10"].to_list())
            if y == primary:
                n_primary = m.height
            else:
                n_fallback += m.height

    combined = pl.concat(frames, how="diagonal_relaxed")

    # How much does the choice of primary vintage matter?
    prim = loads[primary].mapping.select(["hs10", "naics"]).rename({"naics": "_primary"})
    reclass = 0
    examples: list[dict] = []
    for y in ys:
        if y == primary:
            continue
        other = loads[y].mapping.select(["hs10", "naics"]).rename({"naics": "_other"})
        j = prim.join(other, on="hs10", how="inner").filter(
            pl.col("_primary") != pl.col("_other")
        )
        reclass = max(reclass, j.height)
        if j.height and len(examples) < 5:
            for r in j.head(3).iter_rows(named=True):
                examples.append(
                    {
                        "hs10": r["hs10"],
                        f"naics_{primary}": r["_primary"],
                        f"naics_{y}": r["_other"],
                    }
                )

    # Codes the classification revision retired are a different case from
    # reclassification. Census moved this concordance from 2012 NAICS to 2017
    # NAICS with the 2019 vintage, so a 2017-vintage line can carry a NAICS that
    # does not exist in 2017 NAICS at all -- and BEA's 2017 tables are on 2017
    # NAICS. "The primary vintage governs" has no meaning when the primary
    # vintage's answer is not a code in the target classification, so for those
    # lines the revised classification's code for the *same line* is used. This
    # is still one official source; nothing is mapped forward by hand.
    combined, repairs = _repair_retired_naics(combined, loads, ys)

    stack = VintageStack(
        primary_year=primary,
        years=ys,
        mapping=combined,
        n_from_primary=n_primary,
        n_from_fallback=n_fallback,
        n_reclassified_vs_primary=reclass,
        reclassified_examples=examples,
        n_naics_vintage_repaired=sum(r["n_lines"] for r in repairs),
        naics_vintage_repairs=repairs,
    )
    if repairs:
        moved = ", ".join(f"{r['retired_naics']}->{r['successor_naics']}" for r in repairs[:6])
        stack.warnings.append(
            f"{stack.n_naics_vintage_repaired} lines carried a NAICS code retired by the "
            f"2012->2017 classification revision; the successor code for the same line in "
            f"the {max(ys)} vintage was used ({moved})"
        )
    if reclass:
        stack.warnings.append(
            f"{reclass} codes carry a different NAICS in a later vintage than in "
            f"{primary}; the pre-treatment assignment is used, and this count bounds "
            "how much that choice matters"
        )
    if n_fallback:
        stack.warnings.append(
            f"{n_fallback} codes are absent from the {primary} vintage and were taken "
            "from a later one; these are commodity lines created after the primary year"
        )
    return stack
